from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from weekly_business_runtime.errors import MetricStoreError
from weekly_business_runtime.excel_lineage import (
    METADATA_COLUMNS,
    METADATA_WORKSHEET,
    ExcelLineageBinding,
    RevenueExcelLineageAdapter,
)
from weekly_business_runtime.store import StoreBusinessDateReadKey, StoreReadKey


class RevenueExcelLineageAdapterTests(unittest.TestCase):
    def setUp(self) -> None:
        scratch = os.environ.get("PBAC_SCRATCH_ROOT_LOCAL_ONLY")
        self.temp_dir = tempfile.TemporaryDirectory(dir=scratch)
        self.workbook_path = Path(self.temp_dir.name) / "synthetic-revenue-store.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "weekly_revenue"
        sheet.append(("数据当周最后一天", "当周收入", "截止当周最后一天本Q累计收入"))
        sheet.append((20260812, 100, 300))
        workbook.save(self.workbook_path)
        self.adapter = RevenueExcelLineageAdapter()
        self.binding = ExcelLineageBinding(
            store_id="STORE_WEEKLY_REVENUE_HISTORICAL",
            store_asset_id="STORE_ASSET_WEEKLY_REVENUE_SMART_SPEAKER",
            business_context_id="CTX_REVENUE_SMART_SPEAKER_WEEKLY",
            metric_variant_id="MV_REVENUE_SMART_SPEAKER_WEEKLY_EXECUTED_V1",
            workflow_reporting_date="2026-08-13",
            current_revenue_cutoff_date="2026-08-12",
            physical_worksheet="weekly_revenue",
            physical_row=2,
            business_date_column="A",
            result_id="SYNTHETIC-RESULT-1",
            validation_status="passed",
            business_digest="synthetic-digest-1",
        )

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_register_and_exact_reads_preserve_independent_dates(self) -> None:
        self.assertTrue(self.adapter.register(self.workbook_path, self.binding))
        self.assertEqual(
            self.adapter.read_exact(
                self.workbook_path,
                StoreReadKey(
                    self.binding.store_id,
                    self.binding.store_asset_id,
                    self.binding.metric_variant_id,
                    "2026-08-13",
                    self.binding.business_context_id,
                ),
            ),
            self.binding,
        )
        self.assertEqual(
            self.adapter.read_exact_business_date(
                self.workbook_path,
                StoreBusinessDateReadKey(
                    self.binding.store_id,
                    self.binding.store_asset_id,
                    self.binding.metric_variant_id,
                    "2026-08-12",
                    self.binding.business_context_id,
                ),
            ),
            self.binding,
        )
        workbook = load_workbook(self.workbook_path)
        metadata = workbook[METADATA_WORKSHEET]
        self.assertEqual(metadata.sheet_state, "veryHidden")
        self.assertEqual(tuple(cell.value for cell in metadata[1]), METADATA_COLUMNS)
        self.assertNotIn("value", METADATA_COLUMNS)

    def test_idempotent_same_binding_does_not_duplicate_metadata(self) -> None:
        self.adapter.register(self.workbook_path, self.binding)
        self.assertFalse(self.adapter.register(self.workbook_path, self.binding))
        workbook = load_workbook(self.workbook_path)
        self.assertEqual(workbook[METADATA_WORKSHEET].max_row, 2)

    def test_missing_reporting_date_does_not_use_nearest_or_previous_row(self) -> None:
        self.adapter.register(self.workbook_path, self.binding)
        with self.assertRaises(MetricStoreError) as caught:
            self.adapter.read_exact(
                self.workbook_path,
                StoreReadKey(
                    self.binding.store_id,
                    self.binding.store_asset_id,
                    self.binding.metric_variant_id,
                    "2026-08-20",
                    self.binding.business_context_id,
                ),
            )
        self.assertEqual(caught.exception.code, "STORE_EXCEL_LINEAGE_REPORTING_KEY_NOT_FOUND")

    def test_missing_metadata_fails_closed(self) -> None:
        with self.assertRaises(MetricStoreError) as caught:
            self.adapter.read_exact(self.workbook_path, self.binding.reporting_read_key)
        self.assertEqual(caught.exception.code, "STORE_EXCEL_LINEAGE_METADATA_MISSING")

    def test_physical_business_date_mismatch_fails_closed(self) -> None:
        mismatch = ExcelLineageBinding(
            **{**self.binding.__dict__, "current_revenue_cutoff_date": "2026-08-11"}
        )
        with self.assertRaises(MetricStoreError) as caught:
            self.adapter.register(self.workbook_path, mismatch)
        self.assertEqual(
            caught.exception.code, "STORE_EXCEL_LINEAGE_BUSINESS_DATE_MISMATCH"
        )

    def test_duplicate_reporting_key_is_ambiguous(self) -> None:
        self.adapter.register(self.workbook_path, self.binding)
        workbook = load_workbook(self.workbook_path)
        metadata = workbook[METADATA_WORKSHEET]
        duplicate = [cell.value for cell in metadata[2]]
        duplicate[10] = "SYNTHETIC-RESULT-2"
        metadata.append(duplicate)
        workbook.save(self.workbook_path)
        with self.assertRaises(MetricStoreError) as caught:
            self.adapter.read_exact(self.workbook_path, self.binding.reporting_read_key)
        self.assertEqual(
            caught.exception.code, "STORE_EXCEL_LINEAGE_REPORTING_KEY_AMBIGUOUS"
        )


if __name__ == "__main__":
    unittest.main()
