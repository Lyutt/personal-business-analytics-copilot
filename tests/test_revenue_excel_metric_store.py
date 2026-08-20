from __future__ import annotations

import os
import tempfile
import unittest
from dataclasses import replace
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from weekly_business_runtime.errors import MetricStoreError
from weekly_business_runtime.excel_lineage import METADATA_WORKSHEET
from weekly_business_runtime.excel_store import (
    RevenueExcelMetricBinding,
    RevenueExcelMetricStore,
    RevenueExcelStoreAssetConfig,
)
from weekly_business_runtime.store import MetricStoreRecord

STORE_ID = "STORE_WEEKLY_REVENUE_HISTORICAL"
WORKFLOW_ID = "WF_WEEKLY_BUSINESS_REPORT"


class RevenueExcelMetricStoreTests(unittest.TestCase):
    def setUp(self) -> None:
        scratch = os.environ.get("PBAC_SCRATCH_ROOT_LOCAL_ONLY")
        self.temp_dir = tempfile.TemporaryDirectory(dir=scratch)
        root = Path(self.temp_dir.name)
        self.smart_path = root / "synthetic-smart-speaker-store.xlsx"
        self.fast_path = root / "synthetic-fast-version-store.xlsx"
        self.headers = ("business_date", "weekly_revenue", "qtd_revenue")
        self._create_workbook(self.smart_path)
        self._create_workbook(self.fast_path)
        self.smart_config = self._config(
            self.smart_path,
            "STORE_ASSET_WEEKLY_REVENUE_SMART_SPEAKER",
            "CTX_REVENUE_SMART_SPEAKER_WEEKLY",
            "PL_REVENUE_SMART_SPEAKER_WEEKLY",
            "MV_REVENUE_SMART_SPEAKER_WEEKLY_EXECUTED_V1",
            "MV_REVENUE_SMART_SPEAKER_QTD_EXECUTED_V1",
        )
        self.fast_config = self._config(
            self.fast_path,
            "STORE_ASSET_WEEKLY_REVENUE_FAST_VERSION",
            "CTX_REVENUE_FAST_VERSION_WEEKLY",
            "PL_REVENUE_FAST_VERSION_WEEKLY",
            "MV_REVENUE_FAST_VERSION_WEEKLY_EXECUTED_V1",
            "MV_REVENUE_FAST_VERSION_QTD_EXECUTED_V1",
        )
        self.store = RevenueExcelMetricStore(self.smart_config, self.fast_config)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def _create_workbook(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "weekly_revenue"
        worksheet.append(self.headers)
        workbook.save(path)

    def _config(
        self,
        path: Path,
        asset_id: str,
        context_id: str,
        pipeline_id: str,
        weekly_variant: str,
        qtd_variant: str,
    ) -> RevenueExcelStoreAssetConfig:
        return RevenueExcelStoreAssetConfig(
            workbook_path=path,
            store_id=STORE_ID,
            store_asset_id=asset_id,
            business_context_id=context_id,
            workflow_id=WORKFLOW_ID,
            pipeline_id=pipeline_id,
            worksheet_name="weekly_revenue",
            business_date_column="A",
            metric_bindings=(
                RevenueExcelMetricBinding(
                    weekly_variant, "B", "1.0.0-draft", "monetary_amount", "CNY_yuan"
                ),
                RevenueExcelMetricBinding(
                    qtd_variant, "C", "1.0.0-draft", "monetary_amount", "CNY_yuan"
                ),
            ),
            allowed_persisted_variant_sets=(
                frozenset((weekly_variant, qtd_variant)),
                frozenset((qtd_variant,)),
            ),
            expected_headers=(("A", self.headers[0]), ("B", self.headers[1]), ("C", self.headers[2])),
        )

    def _record(
        self,
        config: RevenueExcelStoreAssetConfig,
        variant_id: str,
        value: str,
        *,
        suffix: str,
    ) -> MetricStoreRecord:
        metric = config.metric_binding_by_id[variant_id]
        return MetricStoreRecord(
            result_id=f"SYNTHETIC_RESULT_{suffix}",
            workflow_id=config.workflow_id,
            workflow_run_id="SYNTHETIC_WORKFLOW_RUN",
            pipeline_id=config.pipeline_id,
            pipeline_run_id="SYNTHETIC_PIPELINE_RUN",
            store_id=config.store_id,
            store_asset_id=config.store_asset_id,
            metric_variant_id=variant_id,
            metric_variant_version=metric.metric_variant_version,
            workflow_reporting_date="2026-08-13",
            current_revenue_cutoff_date="2026-08-12",
            business_context_id=config.business_context_id,
            reporting_period="2026-W33",
            value=Decimal(value),
            value_status="valid_value",
            numeric_semantics=metric.numeric_semantics,
            unit=metric.unit,
            precision=metric.precision,
            validation_status="passed",
            generated_at="2026-08-13T00:00:00+08:00",
            lineage_references=("synthetic://dataset",),
        )

    def _regular_records(
        self, config: RevenueExcelStoreAssetConfig
    ) -> tuple[MetricStoreRecord, ...]:
        weekly, qtd = config.metric_bindings
        return (
            self._record(config, weekly.metric_variant_id, "100.25", suffix="WEEKLY"),
            self._record(config, qtd.metric_variant_id, "300.75", suffix="QTD"),
        )

    def test_regular_write_is_atomic_exact_and_verified_for_both_assets(self) -> None:
        for config in (self.smart_config, self.fast_config):
            records = self._regular_records(config)
            plan = self.store.preflight_write(records)
            receipt = self.store.write_validated(plan)
            self.assertTrue(self.store.verify_write(receipt))
            for expected in records:
                self.assertEqual(self.store.read_exact(expected.read_key).value, expected.value)
                self.assertEqual(
                    self.store.read_exact_business_date(expected.business_date_read_key).value,
                    expected.value,
                )
            workbook = load_workbook(config.workbook_path)
            worksheet = workbook[config.worksheet_name]
            self.assertEqual(worksheet["A2"].value.date(), date(2026, 8, 12))
            self.assertEqual(Decimal(str(worksheet["B2"].value)), Decimal("100.25"))
            self.assertEqual(Decimal(str(worksheet["C2"].value)), Decimal("300.75"))
            self.assertEqual(workbook[METADATA_WORKSHEET].sheet_state, "veryHidden")

    def test_idempotent_replay_does_not_append_or_duplicate_metadata(self) -> None:
        records = self._regular_records(self.smart_config)
        first = self.store.write_validated(self.store.preflight_write(records))
        self.assertTrue(self.store.verify_write(first))
        replay_plan = self.store.preflight_write(records)
        self.assertTrue(replay_plan.idempotent_replay)
        replay = self.store.write_validated(replay_plan)
        self.assertTrue(self.store.verify_write(replay))
        workbook = load_workbook(self.smart_path)
        self.assertEqual(workbook["weekly_revenue"].max_row, 2)
        self.assertEqual(workbook[METADATA_WORKSHEET].max_row, 3)

    def test_conflicting_duplicate_is_blocked_without_overwrite(self) -> None:
        records = self._regular_records(self.smart_config)
        receipt = self.store.write_validated(self.store.preflight_write(records))
        self.assertTrue(self.store.verify_write(receipt))
        conflict = (replace(records[0], value=Decimal("999")), records[1])
        with self.assertRaises(MetricStoreError) as caught:
            self.store.preflight_write(conflict)
        self.assertEqual(caught.exception.code, "STORE_DUPLICATE_CONFLICT")
        workbook = load_workbook(self.smart_path)
        self.assertEqual(Decimal(str(workbook["weekly_revenue"]["B2"].value)), Decimal("100.25"))

    def test_quarter_transition_persists_qtd_and_leaves_weekly_blank(self) -> None:
        qtd = self.smart_config.metric_bindings[1]
        records = (self._record(self.smart_config, qtd.metric_variant_id, "40", suffix="QTD"),)
        receipt = self.store.write_validated(self.store.preflight_write(records))
        self.assertTrue(self.store.verify_write(receipt))
        workbook = load_workbook(self.smart_path)
        worksheet = workbook["weekly_revenue"]
        self.assertIsNone(worksheet["B2"].value)
        self.assertEqual(worksheet["C2"].value, 40)

    def test_wow_result_is_not_accepted_as_a_physical_persisted_field(self) -> None:
        regular = self._regular_records(self.smart_config)
        wow = replace(
            regular[0],
            result_id="SYNTHETIC_RESULT_WOW",
            metric_variant_id="MV_REVENUE_SMART_SPEAKER_WEEKLY_EXECUTED_WOW_V1",
            numeric_semantics="ratio",
            unit="decimal_ratio",
            value=Decimal("0.1"),
        )
        with self.assertRaises(MetricStoreError) as caught:
            self.store.preflight_write((*regular, wow))
        self.assertEqual(caught.exception.code, "STORE_WRITE_SET_INCOMPLETE")

    def test_existing_physical_row_without_metadata_fails_closed(self) -> None:
        workbook = load_workbook(self.smart_path)
        workbook["weekly_revenue"].append((date(2026, 8, 12), 100, 300))
        workbook.save(self.smart_path)
        with self.assertRaises(MetricStoreError) as caught:
            self.store.preflight_write(self._regular_records(self.smart_config))
        self.assertEqual(caught.exception.code, "STORE_EXCEL_LINEAGE_METADATA_MISSING")

    def test_physical_value_tamper_breaks_digest_and_post_write_verification(self) -> None:
        records = self._regular_records(self.smart_config)
        receipt = self.store.write_validated(self.store.preflight_write(records))
        workbook = load_workbook(self.smart_path)
        workbook["weekly_revenue"]["B2"] = 999
        workbook.save(self.smart_path)
        self.assertFalse(self.store.verify_write(receipt))
        with self.assertRaises(MetricStoreError) as caught:
            self.store.read_exact(records[0].read_key)
        self.assertEqual(caught.exception.code, "STORE_EXCEL_BUSINESS_DIGEST_MISMATCH")


if __name__ == "__main__":
    unittest.main()
