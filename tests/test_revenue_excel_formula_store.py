from __future__ import annotations

import os
import tempfile
import unittest
import xml.etree.ElementTree as ET
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

from weekly_business_runtime.excel_lineage import (
    ExcelLineageBinding,
    RevenueExcelLineageAdapter,
)
from weekly_business_runtime.excel_store import (
    RevenueExcelMetricBinding,
    RevenueExcelMetricStore,
    RevenueExcelPhysicalSnapshotBinding,
    RevenueExcelStoreAssetConfig,
    _excel_business_digest,
)
from weekly_business_runtime.store import (
    MetricStoreRecord,
    StorePhysicalSnapshotReadKey,
    StorePhysicalValue,
    StoreWriteContext,
)

STORE_ID = "STORE_WEEKLY_REVENUE_HISTORICAL"
WORKFLOW_ID = "WF_WEEKLY_BUSINESS_REPORT"
XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


class SyntheticCachedFormulaEngine:
    """Test-only evaluator that writes cached values while retaining formulas."""

    def __init__(self, profile: str, worksheet_name: str) -> None:
        self.profile = profile
        self.worksheet_name = worksheet_name
        self.called = False

    def recalculate(self, workbook_path: Path) -> None:
        self.called = True
        workbook = load_workbook(workbook_path, data_only=False)
        worksheet = workbook[self.worksheet_name]
        current_row = max(
            row for row in range(2, worksheet.max_row + 1) if worksheet[f"A{row}"].value
        )
        cutoff = worksheet[f"A{current_row}"].value.date()
        target_prior = cutoff.replace(year=cutoff.year - 1) + (date.resolution)
        current_days = Decimal(str((cutoff - self._quarter_start(cutoff)).days + 1))
        prior_days = Decimal(
            str((target_prior - self._quarter_start(target_prior)).days + 1)
        )
        values: dict[str, Decimal] = {}
        if self.profile == "ctv_yoy":
            prior_row = next(
                row
                for row in range(2, current_row)
                if self._cell_date(worksheet[f"A{row}"].value) == target_prior
            )
            values[f"F{current_row}"] = (
                Decimal(str(worksheet[f"B{current_row}"].value)) / current_days
            ) / (Decimal(str(worksheet[f"B{prior_row}"].value)) / prior_days) - 1
        else:
            values[f"F{current_row}"] = (
                Decimal(str(worksheet[f"B{current_row}"].value)) / current_days
            ) / (Decimal(str(worksheet[f"D{current_row}"].value)) / prior_days) - 1
            if worksheet[f"H{current_row}"].value is None:
                values[f"G{current_row}"] = Decimal(
                    str(worksheet[f"B{current_row}"].value)
                )
                values[f"I{current_row}"] = Decimal(
                    str(worksheet[f"D{current_row}"].value)
                )
                values[f"J{current_row}"] = Decimal(
                    str(worksheet[f"E{current_row}"].value)
                )
            else:
                previous_row = max(
                    row
                    for row in range(2, current_row)
                    if worksheet[f"A{row}"].value is not None
                )
                for column in ("B", "C", "D", "E"):
                    target = {"B": "G", "C": "H", "D": "I", "E": "J"}[column]
                    values[f"{target}{current_row}"] = Decimal(
                        str(worksheet[f"{column}{current_row}"].value)
                    ) - Decimal(str(worksheet[f"{column}{previous_row}"].value))
                values[f"K{current_row}"] = (
                    values[f"H{current_row}"]
                    / Decimal(str(worksheet[f"H{previous_row}"].value))
                    - 1
                )
                values[f"L{current_row}"] = (
                    values[f"H{current_row}"] / values[f"J{current_row}"] - 1
                )
        self._write_cached_values(workbook_path, values)

    @staticmethod
    def _quarter_start(value: date) -> date:
        return date(value.year, ((value.month - 1) // 3) * 3 + 1, 1)

    @staticmethod
    def _cell_date(value) -> date:
        return value.date() if hasattr(value, "date") else value

    @staticmethod
    def _write_cached_values(path: Path, values: dict[str, Decimal]) -> None:
        source = BytesIO(path.read_bytes())
        output = BytesIO()
        namespace = {"s": XML_NS}
        with ZipFile(source, "r") as source_zip, ZipFile(
            output, "w", ZIP_DEFLATED
        ) as output_zip:
            for info in source_zip.infolist():
                payload = source_zip.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    root = ET.fromstring(payload)
                    for cell in root.findall(".//s:c", namespace):
                        reference = cell.attrib.get("r")
                        if reference not in values:
                            continue
                        cached = cell.find("s:v", namespace)
                        if cached is None:
                            cached = ET.SubElement(cell, f"{{{XML_NS}}}v")
                        cached.text = format(float(values[reference]), ".15g")
                    payload = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                output_zip.writestr(info, payload)
        path.write_bytes(output.getvalue())


class RevenueExcelFormulaStoreTests(unittest.TestCase):
    def setUp(self) -> None:
        scratch = os.environ.get("PBAC_SCRATCH_ROOT_LOCAL_ONLY")
        self.temp_dir = tempfile.TemporaryDirectory(dir=scratch)
        self.root = Path(self.temp_dir.name)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def _workbook(self, name: str, sheet_name: str, headers: tuple[str, ...]) -> Path:
        path = self.root / name
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        worksheet.append(headers)
        workbook.save(path)
        return path

    @staticmethod
    def _metric(
        variant: str, column: str, semantics: str, unit: str
    ) -> RevenueExcelMetricBinding:
        return RevenueExcelMetricBinding(
            variant, column, "1.0.0-draft", semantics, unit
        )

    @staticmethod
    def _record(
        config: RevenueExcelStoreAssetConfig,
        variant: str,
        value: Decimal,
        *,
        reporting_date: str,
        cutoff_date: str,
    ) -> MetricStoreRecord:
        binding = config.metric_binding_by_id[variant]
        return MetricStoreRecord(
            result_id=f"SYNTH_{reporting_date}_{variant}",
            workflow_id=config.workflow_id,
            workflow_run_id=f"SYNTH_RUN_{reporting_date}",
            pipeline_id=config.pipeline_id,
            pipeline_run_id=f"SYNTH_PIPELINE_{reporting_date}",
            store_id=config.store_id,
            store_asset_id=config.store_asset_id,
            metric_variant_id=variant,
            metric_variant_version=binding.metric_variant_version,
            workflow_reporting_date=reporting_date,
            current_revenue_cutoff_date=cutoff_date,
            business_context_id=config.business_context_id,
            reporting_period=reporting_date,
            value=value,
            value_status="valid_value",
            numeric_semantics=binding.numeric_semantics,
            unit=binding.unit,
            precision=binding.precision,
            validation_status="passed",
            generated_at="2026-08-13T00:00:00+08:00",
            lineage_references=("synthetic://formula-store",),
        )

    @staticmethod
    def _seed_bindings(
        path: Path,
        config: RevenueExcelStoreAssetConfig,
        records: tuple[MetricStoreRecord, ...],
        row: int,
    ) -> None:
        digest = _excel_business_digest(records)
        adapter = RevenueExcelLineageAdapter()
        for record in records:
            adapter.register(
                path,
                ExcelLineageBinding(
                    store_id=record.store_id,
                    store_asset_id=record.store_asset_id,
                    business_context_id=record.business_context_id,
                    metric_variant_id=record.metric_variant_id,
                    workflow_reporting_date=record.workflow_reporting_date,
                    current_revenue_cutoff_date=record.current_revenue_cutoff_date,
                    physical_worksheet=config.worksheet_name,
                    physical_row=row,
                    business_date_column=config.business_date_column,
                    result_id=record.result_id,
                    validation_status="passed",
                    business_digest=digest,
                ),
            )

    def test_ctv_formula_write_uses_exact_prior_row_and_data_only_verification(self) -> None:
        path = self._workbook(
            "synthetic-ctv-formula-store.xlsx",
            "tv",
            ("business_date", "qtd_performance", "qtd_executed", "prior_a", "prior_b", "yoy"),
        )
        variants = (
            "MV_REVENUE_CTV_QTD_PERFORMANCE_V1",
            "MV_REVENUE_CTV_QTD_PERFORMANCE_YOY_V1",
            "MV_REVENUE_CTV_QTD_EXECUTED_V1",
        )
        config = RevenueExcelStoreAssetConfig(
            path,
            STORE_ID,
            "STORE_ASSET_WEEKLY_REVENUE_CTV",
            "CTX_REVENUE_CTV_WEEKLY",
            WORKFLOW_ID,
            "PL_REVENUE_CTV_WEEKLY",
            "tv",
            "A",
            (
                self._metric(variants[0], "B", "monetary_amount", "CNY_yuan"),
                self._metric(variants[1], "F", "ratio", "decimal_ratio"),
                self._metric(variants[2], "C", "monetary_amount", "CNY_yuan"),
            ),
            (frozenset(variants),),
            formula_profile="ctv_yoy",
            formula_columns=("F",),
        )
        workbook = load_workbook(path)
        workbook["tv"].append((date(2025, 8, 13), 100, 70, None, None, None))
        workbook.save(path)
        prior = (
            self._record(
                config,
                variants[0],
                Decimal("100"),
                reporting_date="2025-08-14",
                cutoff_date="2025-08-13",
            ),
        )
        self._seed_bindings(path, config, prior, 2)
        expected_yoy = (Decimal("200") / Decimal("43")) / (
            Decimal("100") / Decimal("44")
        ) - 1
        current = (
            self._record(
                config,
                variants[0],
                Decimal("200"),
                reporting_date="2026-08-13",
                cutoff_date="2026-08-12",
            ),
            self._record(
                config,
                variants[1],
                expected_yoy,
                reporting_date="2026-08-13",
                cutoff_date="2026-08-12",
            ),
            self._record(
                config,
                variants[2],
                Decimal("150"),
                reporting_date="2026-08-13",
                cutoff_date="2026-08-12",
            ),
        )
        engine = SyntheticCachedFormulaEngine("ctv_yoy", "tv")
        store = RevenueExcelMetricStore(config, calculation_engine=engine)
        receipt = store.write_validated(store.preflight_write(current))
        self.assertTrue(engine.called)
        self.assertTrue(store.verify_write(receipt))
        formula_workbook = load_workbook(path, data_only=False)
        value_workbook = load_workbook(path, data_only=True)
        self.assertIsNone(formula_workbook["tv"]["A3"].value)
        self.assertEqual(formula_workbook["tv"]["A4"].value.date(), date(2026, 8, 12))
        self.assertTrue(formula_workbook["tv"]["F4"].value.startswith("="))
        self.assertIsNone(formula_workbook["tv"]["D4"].value)
        self.assertIsNone(formula_workbook["tv"]["E4"].value)
        self.assertAlmostEqual(value_workbook["tv"]["F4"].value, float(expected_yoy))

    def test_technical_regular_formula_write_verifies_all_a_to_l_semantics(self) -> None:
        path = self._workbook(
            "synthetic-technical-formula-store.xlsx",
            "Sheet1",
            tuple(f"field_{column}" for column in "ABCDEFGHIJKL"),
        )
        variants = (
            "MV_REVENUE_TECHNICAL_QTD_PERFORMANCE_V1",
            "MV_REVENUE_TECHNICAL_QTD_PERFORMANCE_YOY_V1",
            "MV_REVENUE_TECHNICAL_QTD_EXECUTED_V1",
            "MV_REVENUE_TECHNICAL_WEEKLY_INCREMENTAL_EXECUTED_V1",
            "MV_REVENUE_TECHNICAL_WEEKLY_INCREMENTAL_EXECUTED_WOW_V1",
            "MV_REVENUE_TECHNICAL_WEEKLY_INCREMENTAL_EXECUTED_YOY_V1",
        )
        columns = ("B", "F", "C", "H", "K", "L")
        semantics = (
            "monetary_amount",
            "ratio",
            "monetary_amount",
            "monetary_amount",
            "ratio",
            "ratio",
        )
        units = (
            "CNY_yuan",
            "decimal_ratio",
            "CNY_yuan",
            "CNY_yuan",
            "decimal_ratio",
            "decimal_ratio",
        )
        config = RevenueExcelStoreAssetConfig(
            path,
            STORE_ID,
            "STORE_ASSET_WEEKLY_REVENUE_TECHNICAL",
            "CTX_REVENUE_TECHNICAL_WEEKLY",
            WORKFLOW_ID,
            "PL_REVENUE_TECHNICAL_WEEKLY",
            "Sheet1",
            "A",
            tuple(
                self._metric(variant, column, semantic, unit)
                for variant, column, semantic, unit in zip(
                    variants, columns, semantics, units, strict=True
                )
            ),
            (
                frozenset(variants),
                frozenset(variants[:3]),
            ),
            formula_profile="technical_full",
            formula_columns=("F", "G", "H", "I", "J", "K", "L"),
            physical_snapshot_bindings=(
                RevenueExcelPhysicalSnapshotBinding(
                    "E",
                    "E",
                    variants[2],
                    variants[2],
                    "prior_year_comparable",
                    "monetary_amount",
                    "CNY_yuan",
                ),
            ),
        )
        workbook = load_workbook(path)
        workbook["Sheet1"].append(
            (date(2026, 8, 5), 180, 120, 90, 100, 0, 0, 30, 0, 20, 0, 0)
        )
        workbook.save(path)
        previous_values = (
            Decimal("180"),
            Decimal("0"),
            Decimal("120"),
            Decimal("30"),
            Decimal("0"),
            Decimal("0"),
        )
        previous = tuple(
            self._record(
                config,
                variant,
                value,
                reporting_date="2026-08-06",
                cutoff_date="2026-08-05",
            )
            for variant, value in zip(variants, previous_values, strict=True)
        )
        self._seed_bindings(path, config, previous, 2)
        snapshot = RevenueExcelMetricStore(config).read_exact_physical_snapshot(
            StorePhysicalSnapshotReadKey(
                STORE_ID,
                config.store_asset_id,
                "E",
                "2026-08-06",
                config.business_context_id,
            )
        )
        self.assertEqual(snapshot.value, Decimal("100"))
        self.assertEqual(snapshot.represented_business_date, "2025-08-06")
        self.assertEqual(snapshot.period_role, "prior_year_comparable")
        expected_yoy = (Decimal("200") / Decimal("43")) / (
            Decimal("100") / Decimal("44")
        ) - 1
        current_values = (
            Decimal("200"),
            expected_yoy,
            Decimal("150"),
            Decimal("30"),
            Decimal("0"),
            Decimal("0.5"),
        )
        current = tuple(
            self._record(
                config,
                variant,
                value,
                reporting_date="2026-08-13",
                cutoff_date="2026-08-12",
            )
            for variant, value in zip(variants, current_values, strict=True)
        )
        context = StoreWriteContext(
            "regular_week",
            (StorePhysicalValue("D", Decimal("100")), StorePhysicalValue("E", Decimal("120"))),
        )
        engine = SyntheticCachedFormulaEngine("technical_full", "Sheet1")
        store = RevenueExcelMetricStore(config, calculation_engine=engine)
        receipt = store.write_validated(store.preflight_write(current, context))
        self.assertTrue(store.verify_write(receipt))
        formula_workbook = load_workbook(path, data_only=False)
        value_workbook = load_workbook(path, data_only=True)
        for column in "FGHIJKL":
            self.assertTrue(formula_workbook["Sheet1"][f"{column}3"].value.startswith("="))
            self.assertIsInstance(value_workbook["Sheet1"][f"{column}3"].value, (int, float))
        self.assertEqual(value_workbook["Sheet1"]["G3"].value, 20)
        self.assertEqual(value_workbook["Sheet1"]["I3"].value, 10)
        self.assertEqual(value_workbook["Sheet1"]["J3"].value, 20)

    def test_technical_quarter_transition_inserts_separator_and_keeps_hkl_blank(self) -> None:
        path = self._workbook(
            "synthetic-technical-quarter-store.xlsx",
            "Sheet1",
            tuple(f"field_{column}" for column in "ABCDEFGHIJKL"),
        )
        base_variants = (
            "MV_REVENUE_TECHNICAL_QTD_PERFORMANCE_V1",
            "MV_REVENUE_TECHNICAL_QTD_PERFORMANCE_YOY_V1",
            "MV_REVENUE_TECHNICAL_QTD_EXECUTED_V1",
        )
        config = RevenueExcelStoreAssetConfig(
            path,
            STORE_ID,
            "STORE_ASSET_WEEKLY_REVENUE_TECHNICAL",
            "CTX_REVENUE_TECHNICAL_WEEKLY",
            WORKFLOW_ID,
            "PL_REVENUE_TECHNICAL_WEEKLY",
            "Sheet1",
            "A",
            (
                self._metric(base_variants[0], "B", "monetary_amount", "CNY_yuan"),
                self._metric(base_variants[1], "F", "ratio", "decimal_ratio"),
                self._metric(base_variants[2], "C", "monetary_amount", "CNY_yuan"),
            ),
            (frozenset(base_variants),),
            formula_profile="technical_full",
            formula_columns=("F", "G", "H", "I", "J", "K", "L"),
        )
        workbook = load_workbook(path)
        workbook["Sheet1"].append(
            (date(2026, 6, 24), 400, 300, 250, 200, 0, 0, 50, 0, 40, 0, 0)
        )
        workbook.save(path)
        expected_yoy = (Decimal("60") / Decimal("1")) / (
            Decimal("30") / Decimal("2")
        ) - 1
        current = tuple(
            self._record(
                config,
                variant,
                value,
                reporting_date="2026-07-02",
                cutoff_date="2026-07-01",
            )
            for variant, value in zip(
                base_variants,
                (Decimal("60"), expected_yoy, Decimal("10")),
                strict=True,
            )
        )
        context = StoreWriteContext(
            "quarter_transition_week",
            (StorePhysicalValue("D", Decimal("30")), StorePhysicalValue("E", Decimal("5"))),
        )
        engine = SyntheticCachedFormulaEngine("technical_full", "Sheet1")
        store = RevenueExcelMetricStore(config, calculation_engine=engine)
        receipt = store.write_validated(store.preflight_write(current, context))
        self.assertTrue(store.verify_write(receipt))
        formula_workbook = load_workbook(path, data_only=False)
        self.assertIsNone(formula_workbook["Sheet1"]["A3"].value)
        self.assertEqual(formula_workbook["Sheet1"]["A4"].value.date(), date(2026, 7, 1))
        for column in ("H", "K", "L"):
            self.assertIsNone(formula_workbook["Sheet1"][f"{column}4"].value)
        for column in ("F", "G", "I", "J"):
            self.assertTrue(formula_workbook["Sheet1"][f"{column}4"].value.startswith("="))

    def test_formula_write_without_calculation_engine_is_not_verified(self) -> None:
        path = self._workbook(
            "synthetic-engine-unavailable.xlsx",
            "tv",
            ("business_date", "qtd_performance", "qtd_executed", "prior_a", "prior_b", "yoy"),
        )
        variants = (
            "MV_REVENUE_CTV_QTD_PERFORMANCE_V1",
            "MV_REVENUE_CTV_QTD_PERFORMANCE_YOY_V1",
            "MV_REVENUE_CTV_QTD_EXECUTED_V1",
        )
        config = RevenueExcelStoreAssetConfig(
            path,
            STORE_ID,
            "STORE_ASSET_WEEKLY_REVENUE_CTV",
            "CTX_REVENUE_CTV_WEEKLY",
            WORKFLOW_ID,
            "PL_REVENUE_CTV_WEEKLY",
            "tv",
            "A",
            (
                self._metric(variants[0], "B", "monetary_amount", "CNY_yuan"),
                self._metric(variants[1], "F", "ratio", "decimal_ratio"),
                self._metric(variants[2], "C", "monetary_amount", "CNY_yuan"),
            ),
            (frozenset(variants),),
            formula_profile="ctv_yoy",
            formula_columns=("F",),
        )
        workbook = load_workbook(path)
        workbook["tv"].append((date(2025, 8, 13), 100, 70, None, None, None))
        workbook.save(path)
        prior = (
            self._record(
                config,
                variants[0],
                Decimal("100"),
                reporting_date="2025-08-14",
                cutoff_date="2025-08-13",
            ),
        )
        self._seed_bindings(path, config, prior, 2)
        yoy = (Decimal("200") / Decimal("43")) / (Decimal("100") / Decimal("44")) - 1
        records = tuple(
            self._record(
                config,
                variant,
                value,
                reporting_date="2026-08-13",
                cutoff_date="2026-08-12",
            )
            for variant, value in zip(
                variants, (Decimal("200"), yoy, Decimal("150")), strict=True
            )
        )
        store = RevenueExcelMetricStore(config)
        receipt = store.write_validated(store.preflight_write(records))
        self.assertFalse(store.verify_write(receipt))


if __name__ == "__main__":
    unittest.main()
