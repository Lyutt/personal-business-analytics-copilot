"""MetricStorePort implementation for registered Revenue Excel Store assets."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass, replace
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.workbook import Workbook

from .errors import MetricStoreError
from .excel_calculation import WorkbookCalculationEngine
from .excel_lineage import ExcelLineageBinding, RevenueExcelLineageAdapter
from .store import (
    MetricStoreRecord,
    StoreBusinessDateReadKey,
    StorePhysicalSnapshot,
    StorePhysicalSnapshotReadKey,
    StoreReadKey,
    StoreWriteContext,
    StoreWriteIdentity,
    StoreWritePlan,
    StoreWriteReceipt,
    _business_digest,
)


@dataclass(frozen=True)
class RevenueExcelMetricBinding:
    """One persisted Metric Variant and its physical static-value column."""

    metric_variant_id: str
    column: str
    metric_variant_version: str
    numeric_semantics: str
    unit: str
    precision: str = "preserve_source_precision"


@dataclass(frozen=True)
class RevenueExcelPhysicalSnapshotBinding:
    """Registered adapter-only read binding for one physical helper snapshot."""

    field_id: str
    column: str
    anchor_metric_variant_id: str
    metric_variant_id: str
    period_role: str
    numeric_semantics: str
    unit: str


@dataclass(frozen=True)
class RevenueExcelStoreAssetConfig:
    """Adapter configuration backed only by registered Store/Pipeline contracts."""

    workbook_path: Path
    store_id: str
    store_asset_id: str
    business_context_id: str
    workflow_id: str
    pipeline_id: str
    worksheet_name: str
    business_date_column: str
    metric_bindings: tuple[RevenueExcelMetricBinding, ...]
    allowed_persisted_variant_sets: tuple[frozenset[str], ...]
    expected_headers: tuple[tuple[str, str], ...] = ()
    formula_profile: str | None = None
    formula_columns: tuple[str, ...] = ()
    physical_snapshot_bindings: tuple[RevenueExcelPhysicalSnapshotBinding, ...] = ()

    @property
    def metric_binding_by_id(self) -> dict[str, RevenueExcelMetricBinding]:
        return {item.metric_variant_id: item for item in self.metric_bindings}

    @property
    def physical_snapshot_binding_by_id(
        self,
    ) -> dict[str, RevenueExcelPhysicalSnapshotBinding]:
        return {item.field_id: item for item in self.physical_snapshot_bindings}


def _canonical_date(value: object, *, field_name: str, workbook_epoch=None) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and workbook_epoch is not None:
        try:
            converted = from_excel(value, workbook_epoch)
        except (TypeError, ValueError, OverflowError):
            converted = None
        if isinstance(converted, datetime):
            return converted.date().isoformat()
        if isinstance(converted, date):
            return converted.isoformat()
    text = str(value).strip()
    if len(text) == 8 and text.isdigit():
        text = f"{text[:4]}-{text[4:6]}-{text[6:]}"
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError as exc:
        raise MetricStoreError(
            "STORE_EXCEL_BUSINESS_DATE_INVALID",
            f"{field_name} is not an exact supported business date",
        ) from exc


def _decimal_value(value: object, *, cell_reference: str) -> Decimal:
    if value is None or isinstance(value, bool):
        raise MetricStoreError(
            "STORE_EXCEL_VALUE_NOT_CONSUMABLE",
            f"{cell_reference} does not contain a persisted numeric value",
        )
    if isinstance(value, str) and value.startswith("="):
        raise MetricStoreError(
            "STORE_EXCEL_FORMULA_ADAPTER_REQUIRED",
            f"{cell_reference} is formula-backed and requires the formula-capable adapter",
        )
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise MetricStoreError(
            "STORE_EXCEL_VALUE_NOT_CONSUMABLE",
            f"{cell_reference} does not contain a persisted numeric value",
        ) from exc


def _excel_decimal(value: Decimal) -> Decimal:
    """Canonicalize to the 15 significant digits preserved by an Excel cell."""

    return Decimal(format(float(value), ".15g"))


def _excel_business_digest(records: tuple[MetricStoreRecord, ...]) -> str:
    normalized = tuple(replace(item, value=_excel_decimal(item.value)) for item in records)
    return _business_digest(normalized)


class RevenueExcelMetricStore:
    """Exact-key Excel adapter with formula-engine post-write verification."""

    def __init__(
        self,
        *assets: RevenueExcelStoreAssetConfig,
        calculation_engine: WorkbookCalculationEngine | None = None,
    ) -> None:
        if not assets:
            raise ValueError("At least one Revenue Excel Store Asset is required")
        self._assets = {item.store_asset_id: item for item in assets}
        if len(self._assets) != len(assets):
            raise ValueError("Revenue Excel Store Asset identities must be unique")
        self._lineage = RevenueExcelLineageAdapter()
        self._calculation_engine = calculation_engine
        self._calculation_failures: set[StoreWriteIdentity] = set()
        self._validate_configs()

    def read_exact(self, key: StoreReadKey) -> MetricStoreRecord:
        config = self._config_for_key(key.store_id, key.store_asset_id, key.business_context_id)
        binding = self._lineage.read_exact(config.workbook_path, key)
        return self._verified_record_group(config, binding, key.metric_variant_id)

    def read_exact_business_date(
        self, key: StoreBusinessDateReadKey
    ) -> MetricStoreRecord:
        config = self._config_for_key(key.store_id, key.store_asset_id, key.business_context_id)
        binding = self._lineage.read_exact_business_date(config.workbook_path, key)
        return self._verified_record_group(config, binding, key.metric_variant_id)

    def read_exact_physical_snapshot(
        self, key: StorePhysicalSnapshotReadKey
    ) -> StorePhysicalSnapshot:
        config = self._config_for_key(
            key.store_id, key.store_asset_id, key.business_context_id
        )
        snapshot_binding = config.physical_snapshot_binding_by_id.get(key.field_id)
        if snapshot_binding is None:
            raise MetricStoreError(
                "STORE_EXCEL_PHYSICAL_SNAPSHOT_BINDING_MISSING",
                "No registered physical snapshot binding matches the requested field",
            )
        anchor_key = StoreReadKey(
            key.store_id,
            key.store_asset_id,
            snapshot_binding.anchor_metric_variant_id,
            key.workflow_reporting_date,
            key.business_context_id,
        )
        lineage_binding = self._lineage.read_exact(config.workbook_path, anchor_key)
        self._verified_record_group(
            config, lineage_binding, snapshot_binding.anchor_metric_variant_id
        )
        value_workbook = self._load_and_validate(config, data_only=True)
        cell_reference = f"{snapshot_binding.column}{lineage_binding.physical_row}"
        value = _decimal_value(
            value_workbook[config.worksheet_name][cell_reference].value,
            cell_reference=cell_reference,
        )
        represented_business_date = self._prior_year_business_date(
            lineage_binding.current_revenue_cutoff_date
        )
        return StorePhysicalSnapshot(
            read_key=key,
            metric_variant_id=snapshot_binding.metric_variant_id,
            period_role=snapshot_binding.period_role,
            represented_business_date=represented_business_date,
            value=value,
            numeric_semantics=snapshot_binding.numeric_semantics,
            unit=snapshot_binding.unit,
            validation_status=lineage_binding.validation_status,
            lineage_references=(
                f"metric-store-excel-physical://{key.store_asset_id}/"
                f"{config.worksheet_name}#row={lineage_binding.physical_row}"
                f"&column={snapshot_binding.column}",
            ),
        )

    def preflight_write(
        self,
        records: tuple[MetricStoreRecord, ...],
        physical_write_context: StoreWriteContext | None = None,
    ) -> StoreWritePlan:
        if not records:
            raise MetricStoreError(
                "STORE_WRITE_SET_INCOMPLETE", "Validated persisted Store write set is empty"
            )
        identities = {item.write_identity for item in records}
        if len(identities) != 1:
            raise MetricStoreError(
                "STORE_WRITE_IDENTITY_MISMATCH",
                "Every persisted record must share one Revenue business identity",
            )
        if len({item.metric_variant_id for item in records}) != len(records):
            raise MetricStoreError(
                "STORE_WRITE_SET_DUPLICATE_METRIC",
                "Persisted Store write set contains duplicate Metric identities",
            )
        identity = next(iter(identities))
        config = self._config_for_key(
            identity.store_id, identity.store_asset_id, identity.business_context_id
        )
        self._validate_record_set(config, records)
        self._validate_physical_write_context(
            config, records, physical_write_context
        )
        digest = _excel_business_digest(records)
        workbook = self._load_and_validate(config)
        physical_rows = self._physical_rows_for_date(
            workbook, config, identity.current_revenue_cutoff_date
        )
        existing_bindings = self._bindings_for_identity(workbook, identity)
        if not physical_rows and existing_bindings:
            raise MetricStoreError(
                "STORE_EXCEL_LINEAGE_PHYSICAL_TARGET_INVALID",
                "Technical metadata exists without its physical business row",
            )
        if physical_rows and not existing_bindings:
            raise MetricStoreError(
                "STORE_EXCEL_LINEAGE_METADATA_MISSING",
                "The business date exists without required Adapter technical metadata",
            )
        if len(physical_rows) > 1:
            raise MetricStoreError(
                "STORE_EXCEL_BUSINESS_DATE_AMBIGUOUS",
                "More than one physical row has the exact Revenue business date",
            )
        if existing_bindings:
            existing = self._records_from_bindings(config, workbook, existing_bindings)
            if (
                {item.metric_variant_id for item in existing}
                == {item.metric_variant_id for item in records}
                and _excel_business_digest(existing) == digest
                and self._physical_context_matches(
                    config,
                    workbook,
                    physical_rows[0],
                    records,
                    physical_write_context,
                )
            ):
                return StoreWritePlan(
                    identity, records, digest, True, physical_write_context
                )
            raise MetricStoreError(
                "STORE_DUPLICATE_CONFLICT",
                "Revenue business date already contains a different persisted result set",
            )
        return StoreWritePlan(
            identity, records, digest, False, physical_write_context
        )

    def write_validated(self, plan: StoreWritePlan) -> StoreWriteReceipt:
        config = self._config_for_key(
            plan.write_identity.store_id,
            plan.write_identity.store_asset_id,
            plan.write_identity.business_context_id,
        )
        if _excel_business_digest(plan.records) != plan.business_digest:
            raise MetricStoreError(
                "STORE_WRITE_PLAN_INVALID", "Store write plan business semantics changed"
            )
        if plan.idempotent_replay:
            records = self._records_for_identity(config, plan.write_identity)
            return self._receipt(plan, records)

        self._assert_workbook_write_available(config)
        workbook = self._load_and_validate(config)
        if self._physical_rows_for_date(
            workbook, config, plan.write_identity.current_revenue_cutoff_date
        ):
            raise MetricStoreError(
                "STORE_DUPLICATE_CONFLICT",
                "Revenue business date appeared after preflight; write was not attempted",
            )
        worksheet = workbook[config.worksheet_name]
        physical_row = self._next_physical_row(
            workbook,
            config,
            plan.write_identity.current_revenue_cutoff_date,
        )
        self._copy_previous_row_style(worksheet, physical_row, config)
        worksheet[f"{config.business_date_column}{physical_row}"] = date.fromisoformat(
            plan.write_identity.current_revenue_cutoff_date
        )
        physical_values = {
            item.field_id: item.value
            for item in (
                plan.physical_write_context.physical_values
                if plan.physical_write_context is not None
                else ()
            )
        }
        for column, value in physical_values.items():
            worksheet[f"{column}{physical_row}"] = float(value)
        bindings = config.metric_binding_by_id
        for record in plan.records:
            metric = bindings[record.metric_variant_id]
            if metric.column not in config.formula_columns:
                worksheet[f"{metric.column}{physical_row}"] = float(record.value)
            lineage = ExcelLineageBinding(
                store_id=record.store_id,
                store_asset_id=record.store_asset_id,
                business_context_id=record.business_context_id,
                metric_variant_id=record.metric_variant_id,
                workflow_reporting_date=record.workflow_reporting_date,
                current_revenue_cutoff_date=record.current_revenue_cutoff_date,
                physical_worksheet=config.worksheet_name,
                physical_row=physical_row,
                business_date_column=config.business_date_column,
                result_id=record.result_id,
                validation_status=record.validation_status,
                business_digest=plan.business_digest,
            )
            self._lineage.register_in_workbook(workbook, lineage)
        self._write_formula_profile(
            workbook,
            config,
            plan,
            physical_row,
        )
        try:
            workbook.save(config.workbook_path)
        except (OSError, PermissionError) as exc:
            raise MetricStoreError(
                "STORE_EXCEL_SAVE_FAILED",
                "Revenue Store workbook could not be saved",
            ) from exc
        if config.formula_profile is not None:
            if self._calculation_engine is None:
                self._calculation_failures.add(plan.write_identity)
            else:
                try:
                    self._calculation_engine.recalculate(config.workbook_path)
                except MetricStoreError:
                    self._calculation_failures.add(plan.write_identity)
        return self._receipt(plan, plan.records)

    def verify_write(self, receipt: StoreWriteReceipt) -> bool:
        if receipt.write_identity in self._calculation_failures:
            return False
        try:
            config = self._config_for_key(
                receipt.write_identity.store_id,
                receipt.write_identity.store_asset_id,
                receipt.write_identity.business_context_id,
            )
            records = self._records_for_identity(config, receipt.write_identity)
            formula_verified = self._verify_formula_profile(
                config, receipt.write_identity
            )
        except MetricStoreError:
            return False
        return (
            tuple(item.read_key for item in records) == receipt.read_keys
            and tuple(item.result_id for item in records) == receipt.result_ids
            and _excel_business_digest(records) == receipt.business_digest
            and formula_verified
        )

    def _validate_configs(self) -> None:
        for config in self._assets.values():
            variants = [item.metric_variant_id for item in config.metric_bindings]
            columns = [item.column for item in config.metric_bindings]
            if len(set(variants)) != len(variants) or len(set(columns)) != len(columns):
                raise ValueError("Metric Variant and physical value columns must be unique")
            registered = frozenset(variants)
            if not config.allowed_persisted_variant_sets or any(
                not allowed or not allowed <= registered
                for allowed in config.allowed_persisted_variant_sets
            ):
                raise ValueError("Allowed persisted sets must be non-empty registered subsets")
            if config.formula_profile not in {None, "ctv_yoy", "technical_full"}:
                raise ValueError("Revenue Excel formula profile is not registered")
            if config.formula_profile is None and config.formula_columns:
                raise ValueError("Static Store Assets cannot declare formula columns")
            if config.formula_profile is not None and not config.formula_columns:
                raise ValueError("Formula-backed Store Assets require formula columns")
            expected_formula_columns = {
                "ctv_yoy": ("F",),
                "technical_full": ("F", "G", "H", "I", "J", "K", "L"),
            }
            if config.formula_profile is not None and config.formula_columns != (
                expected_formula_columns[config.formula_profile]
            ):
                raise ValueError(
                    "Formula-backed Store Asset columns must match the registered profile"
                )
            snapshots = config.physical_snapshot_bindings
            if len({item.field_id for item in snapshots}) != len(snapshots):
                raise ValueError("Physical snapshot field identities must be unique")
            if snapshots and config.formula_profile != "technical_full":
                raise ValueError(
                    "Only the registered Technical formula profile exposes physical snapshots"
                )
            if any(
                item.anchor_metric_variant_id not in registered
                or item.metric_variant_id not in registered
                or item.period_role != "prior_year_comparable"
                or item.column != item.field_id
                for item in snapshots
            ):
                raise ValueError(
                    "Physical snapshot bindings must use registered exact Technical lineage"
                )

    def _validate_record_set(
        self, config: RevenueExcelStoreAssetConfig, records: tuple[MetricStoreRecord, ...]
    ) -> None:
        variant_set = frozenset(item.metric_variant_id for item in records)
        if variant_set not in config.allowed_persisted_variant_sets:
            raise MetricStoreError(
                "STORE_WRITE_SET_INCOMPLETE",
                "Persisted Metric Variant set does not match an authoritative report-mode set",
            )
        bindings = config.metric_binding_by_id
        for record in records:
            metric = bindings[record.metric_variant_id]
            if (
                record.workflow_id != config.workflow_id
                or record.pipeline_id != config.pipeline_id
                or record.metric_variant_version != metric.metric_variant_version
                or record.validation_status != "passed"
                or record.value_status != "valid_value"
                or record.numeric_semantics != metric.numeric_semantics
                or record.unit != metric.unit
                or record.precision != metric.precision
                or not isinstance(record.value, Decimal)
                or not record.value.is_finite()
            ):
                raise MetricStoreError(
                    "STORE_WRITE_REQUIRES_VALIDATED_RESULT",
                    "Persisted record does not satisfy its registered Metric Store binding",
                )

    @staticmethod
    def _validate_physical_write_context(
        config: RevenueExcelStoreAssetConfig,
        records: tuple[MetricStoreRecord, ...],
        context: StoreWriteContext | None,
    ) -> None:
        if config.formula_profile == "technical_full":
            if context is None or context.report_mode not in {
                "regular_week",
                "quarter_transition_week",
            }:
                raise MetricStoreError(
                    "STORE_EXCEL_PHYSICAL_CONTEXT_MISSING",
                    "Technical Store formula write requires an exact report mode",
                )
            values = {item.field_id: item.value for item in context.physical_values}
            if len(values) != len(context.physical_values) or set(values) != {"D", "E"}:
                raise MetricStoreError(
                    "STORE_EXCEL_PHYSICAL_CONTEXT_MISSING",
                    "Technical Store formula write requires unique D and E values",
                )
            if any(
                not isinstance(value, Decimal) or not value.is_finite()
                for value in values.values()
            ):
                raise MetricStoreError(
                    "STORE_EXCEL_PHYSICAL_CONTEXT_MISSING",
                    "Technical Store formula write requires finite Decimal D and E values",
                )
            persisted_columns = {
                config.metric_binding_by_id[item.metric_variant_id].column
                for item in records
            }
            weekly_columns = {"H", "K", "L"}
            present_weekly_columns = persisted_columns & weekly_columns
            if present_weekly_columns not in (set(), weekly_columns):
                raise MetricStoreError(
                    "STORE_WRITE_SET_INCOMPLETE",
                    "Technical weekly Metric columns must be complete or not applicable",
                )
            weekly_columns_present = present_weekly_columns == weekly_columns
            if (context.report_mode == "regular_week") != weekly_columns_present:
                raise MetricStoreError(
                    "STORE_WRITE_SET_INCOMPLETE",
                    "Technical persisted Metric set does not match its report mode",
                )
        elif context is not None and context.physical_values:
            raise MetricStoreError(
                "STORE_EXCEL_PHYSICAL_CONTEXT_UNEXPECTED",
                "This Revenue Store Asset does not accept supplemental physical values",
            )

    def _config_for_key(
        self, store_id: str, store_asset_id: str, business_context_id: str
    ) -> RevenueExcelStoreAssetConfig:
        config = self._assets.get(store_asset_id)
        if config is None or (
            config.store_id != store_id or config.business_context_id != business_context_id
        ):
            raise MetricStoreError(
                "STORE_EXCEL_ASSET_NOT_CONFIGURED",
                "No exact Revenue Excel Store Asset configuration matches the key",
            )
        return config

    def _load_and_validate(
        self, config: RevenueExcelStoreAssetConfig, *, data_only: bool = False
    ) -> Workbook:
        if not config.workbook_path.is_file():
            raise MetricStoreError(
                "STORE_EXCEL_WORKBOOK_UNAVAILABLE",
                "Configured Revenue Store workbook is unavailable",
            )
        workbook = load_workbook(config.workbook_path, data_only=data_only)
        if config.worksheet_name not in workbook.sheetnames:
            raise MetricStoreError(
                "STORE_EXCEL_WORKSHEET_MISSING",
                "Configured Revenue Store worksheet is unavailable",
            )
        worksheet = workbook[config.worksheet_name]
        for column, expected in config.expected_headers:
            if worksheet[f"{column}1"].value != expected:
                raise MetricStoreError(
                    "STORE_EXCEL_HEADER_MISMATCH",
                    f"Revenue Store header {column} does not match its configured contract",
                )
        return workbook

    @staticmethod
    def _assert_workbook_write_available(config: RevenueExcelStoreAssetConfig) -> None:
        if not config.workbook_path.is_file():
            raise MetricStoreError(
                "STORE_EXCEL_WORKBOOK_UNAVAILABLE",
                "Configured Revenue Store workbook is unavailable",
            )
        lock_file = config.workbook_path.with_name(f"~${config.workbook_path.name}")
        if lock_file.exists():
            raise MetricStoreError(
                "STORE_EXCEL_WORKBOOK_LOCKED",
                "Configured Revenue Store workbook is occupied",
            )
        try:
            with config.workbook_path.open("r+b"):
                pass
        except (OSError, PermissionError) as exc:
            raise MetricStoreError(
                "STORE_EXCEL_WORKBOOK_LOCKED",
                "Configured Revenue Store workbook is not available for writing",
            ) from exc

    def _physical_context_matches(
        self,
        config: RevenueExcelStoreAssetConfig,
        workbook: Workbook,
        physical_row: int,
        records: tuple[MetricStoreRecord, ...],
        context: StoreWriteContext | None,
    ) -> bool:
        if config.formula_profile != "technical_full":
            return context is None or not context.physical_values
        if context is None:
            return False
        persisted_columns = {
            config.metric_binding_by_id[item.metric_variant_id].column for item in records
        }
        expected_mode = (
            "regular_week"
            if {"H", "K", "L"}.issubset(persisted_columns)
            else "quarter_transition_week"
        )
        if context.report_mode != expected_mode:
            return False
        expected_values = {item.field_id: item.value for item in context.physical_values}
        worksheet = workbook[config.worksheet_name]
        try:
            actual_values = {
                column: _decimal_value(
                    worksheet[f"{column}{physical_row}"].value,
                    cell_reference=f"{column}{physical_row}",
                )
                for column in ("D", "E")
            }
        except MetricStoreError:
            return False
        return all(
            _excel_decimal(actual_values[column])
            == _excel_decimal(expected_values[column])
            for column in ("D", "E")
        )

    def _physical_rows_for_date(
        self, workbook: Workbook, config: RevenueExcelStoreAssetConfig, business_date: str
    ) -> tuple[int, ...]:
        worksheet = workbook[config.worksheet_name]
        matches: list[int] = []
        for row in range(2, worksheet.max_row + 1):
            value = worksheet[f"{config.business_date_column}{row}"].value
            if value is None:
                continue
            try:
                candidate = _canonical_date(
                    value, field_name="physical business-date cell", workbook_epoch=workbook.epoch
                )
            except MetricStoreError:
                continue
            if candidate == business_date:
                matches.append(row)
        return tuple(matches)

    def _bindings_for_identity(
        self, workbook: Workbook, identity: StoreWriteIdentity
    ) -> tuple[ExcelLineageBinding, ...]:
        try:
            bindings = self._lineage.bindings_in_workbook(workbook)
        except MetricStoreError as exc:
            if exc.code == "STORE_EXCEL_LINEAGE_METADATA_MISSING":
                return ()
            raise
        return tuple(
            item
            for item in bindings
            if item.store_id == identity.store_id
            and item.store_asset_id == identity.store_asset_id
            and item.business_context_id == identity.business_context_id
            and item.current_revenue_cutoff_date == identity.current_revenue_cutoff_date
        )

    def _verified_record_group(
        self,
        config: RevenueExcelStoreAssetConfig,
        target: ExcelLineageBinding,
        metric_variant_id: str,
    ) -> MetricStoreRecord:
        workbook = self._load_and_validate(config)
        all_bindings = self._lineage.bindings_in_workbook(workbook)
        group = tuple(
            item
            for item in all_bindings
            if item.store_id == target.store_id
            and item.store_asset_id == target.store_asset_id
            and item.business_context_id == target.business_context_id
            and item.workflow_reporting_date == target.workflow_reporting_date
            and item.current_revenue_cutoff_date == target.current_revenue_cutoff_date
            and item.physical_worksheet == target.physical_worksheet
            and item.physical_row == target.physical_row
            and item.business_digest == target.business_digest
        )
        records = self._records_from_bindings(config, workbook, group)
        if not records or _excel_business_digest(records) != target.business_digest:
            raise MetricStoreError(
                "STORE_EXCEL_BUSINESS_DIGEST_MISMATCH",
                "Physical Revenue row does not match its validated result-set digest",
            )
        matches = [item for item in records if item.metric_variant_id == metric_variant_id]
        if len(matches) != 1:
            raise MetricStoreError(
                "STORE_EXCEL_LINEAGE_REPORTING_KEY_AMBIGUOUS",
                "Validated result group does not contain exactly one requested Metric Variant",
            )
        return matches[0]

    def _records_for_identity(
        self, config: RevenueExcelStoreAssetConfig, identity: StoreWriteIdentity
    ) -> tuple[MetricStoreRecord, ...]:
        workbook = self._load_and_validate(config)
        bindings = self._bindings_for_identity(workbook, identity)
        records = self._records_from_bindings(config, workbook, bindings)
        if not bindings or len({item.business_digest for item in bindings}) != 1:
            raise MetricStoreError(
                "STORE_EXCEL_LINEAGE_BUSINESS_DATE_KEY_AMBIGUOUS",
                "Revenue business identity does not have one complete metadata group",
            )
        digest = bindings[0].business_digest
        if _excel_business_digest(records) != digest:
            raise MetricStoreError(
                "STORE_EXCEL_BUSINESS_DIGEST_MISMATCH",
                "Physical Revenue row does not match its validated result-set digest",
            )
        return records

    def _records_from_bindings(
        self,
        config: RevenueExcelStoreAssetConfig,
        workbook: Workbook,
        bindings: tuple[ExcelLineageBinding, ...],
    ) -> tuple[MetricStoreRecord, ...]:
        value_workbook = (
            self._load_and_validate(config, data_only=True)
            if config.formula_profile is not None
            else workbook
        )
        return tuple(
            sorted(
                (
                    self._record_from_binding(config, value_workbook, item)
                    for item in bindings
                ),
                key=lambda item: item.metric_variant_id,
            )
        )

    def _record_from_binding(
        self,
        config: RevenueExcelStoreAssetConfig,
        workbook: Workbook,
        binding: ExcelLineageBinding,
    ) -> MetricStoreRecord:
        metric = config.metric_binding_by_id.get(binding.metric_variant_id)
        if metric is None:
            raise MetricStoreError(
                "STORE_EXCEL_METRIC_BINDING_MISSING",
                "Technical lineage references a Metric Variant not persisted by this Store Asset",
            )
        worksheet = workbook[binding.physical_worksheet]
        cell_reference = f"{metric.column}{binding.physical_row}"
        value = _decimal_value(worksheet[cell_reference].value, cell_reference=cell_reference)
        return MetricStoreRecord(
            result_id=binding.result_id,
            workflow_id=config.workflow_id,
            workflow_run_id=f"historical:{binding.workflow_reporting_date}",
            pipeline_id=config.pipeline_id,
            pipeline_run_id=f"historical:{binding.workflow_reporting_date}",
            store_id=binding.store_id,
            store_asset_id=binding.store_asset_id,
            metric_variant_id=binding.metric_variant_id,
            metric_variant_version=metric.metric_variant_version,
            workflow_reporting_date=binding.workflow_reporting_date,
            current_revenue_cutoff_date=binding.current_revenue_cutoff_date,
            business_context_id=binding.business_context_id,
            reporting_period=binding.workflow_reporting_date,
            value=value,
            value_status="valid_value",
            numeric_semantics=metric.numeric_semantics,
            unit=metric.unit,
            precision=metric.precision,
            validation_status=binding.validation_status,
            generated_at="",
            lineage_references=(
                f"metric-store-excel://{binding.store_asset_id}/{binding.physical_worksheet}"
                f"#row={binding.physical_row}",
            ),
        )

    def _next_physical_row(
        self,
        workbook: Workbook,
        config: RevenueExcelStoreAssetConfig,
        business_date: str,
    ) -> int:
        worksheet = workbook[config.worksheet_name]
        dated_rows: list[tuple[int, date]] = []
        for row in range(2, worksheet.max_row + 1):
            raw = worksheet[f"{config.business_date_column}{row}"].value
            if raw is None:
                continue
            try:
                parsed = date.fromisoformat(
                    _canonical_date(
                        raw,
                        field_name="physical business-date cell",
                        workbook_epoch=workbook.epoch,
                    )
                )
            except MetricStoreError:
                continue
            dated_rows.append((row, parsed))
        if not dated_rows:
            return 2
        last_row, last_date = max(dated_rows, key=lambda item: item[0])
        target_date = date.fromisoformat(business_date)
        last_quarter = (last_date.year, (last_date.month - 1) // 3)
        target_quarter = (target_date.year, (target_date.month - 1) // 3)
        return last_row + (2 if last_quarter != target_quarter else 1)

    def _write_formula_profile(
        self,
        workbook: Workbook,
        config: RevenueExcelStoreAssetConfig,
        plan: StoreWritePlan,
        physical_row: int,
    ) -> None:
        if config.formula_profile is None:
            return
        worksheet = workbook[config.worksheet_name]
        if config.formula_profile == "ctv_yoy":
            prior_date = self._prior_year_business_date(
                plan.write_identity.current_revenue_cutoff_date
            )
            prior_row = self._exact_business_row(
                workbook,
                config,
                prior_date,
                self._metric_variant_for_column(config, "B"),
            )
            worksheet[f"F{physical_row}"] = self._day_normalized_yoy_formula(
                physical_row, f"B{prior_row}"
            )
            return

        context = plan.physical_write_context
        if context is None:
            raise MetricStoreError(
                "STORE_EXCEL_PHYSICAL_CONTEXT_MISSING",
                "Technical formula profile requires physical write context",
            )
        worksheet[f"F{physical_row}"] = self._day_normalized_yoy_formula(
            physical_row, f"D{physical_row}"
        )
        if context.report_mode == "quarter_transition_week":
            worksheet[f"G{physical_row}"] = f"=B{physical_row}"
            worksheet[f"H{physical_row}"] = None
            worksheet[f"I{physical_row}"] = f"=D{physical_row}"
            worksheet[f"J{physical_row}"] = f"=E{physical_row}"
            worksheet[f"K{physical_row}"] = None
            worksheet[f"L{physical_row}"] = None
            return
        reporting_date = date.fromisoformat(plan.records[0].workflow_reporting_date)
        previous_row = self._exact_reporting_row(
            workbook,
            config,
            (reporting_date - timedelta(days=7)).isoformat(),
        )
        worksheet[f"G{physical_row}"] = f"=B{physical_row}-B{previous_row}"
        worksheet[f"H{physical_row}"] = f"=C{physical_row}-C{previous_row}"
        worksheet[f"I{physical_row}"] = f"=D{physical_row}-D{previous_row}"
        worksheet[f"J{physical_row}"] = f"=E{physical_row}-E{previous_row}"
        worksheet[f"K{physical_row}"] = f"=H{physical_row}/H{previous_row}-1"
        worksheet[f"L{physical_row}"] = f"=H{physical_row}/J{physical_row}-1"

    def _verify_formula_profile(
        self,
        config: RevenueExcelStoreAssetConfig,
        identity: StoreWriteIdentity,
    ) -> bool:
        if config.formula_profile is None:
            return True
        formula_workbook = self._load_and_validate(config)
        value_workbook = self._load_and_validate(config, data_only=True)
        rows = self._physical_rows_for_date(
            formula_workbook, config, identity.current_revenue_cutoff_date
        )
        if len(rows) != 1:
            return False
        row = rows[0]
        formula_sheet = formula_workbook[config.worksheet_name]
        value_sheet = value_workbook[config.worksheet_name]
        if config.formula_profile == "ctv_yoy":
            prior_date = self._prior_year_business_date(
                identity.current_revenue_cutoff_date
            )
            prior_row = self._exact_business_row(
                formula_workbook,
                config,
                prior_date,
                self._metric_variant_for_column(config, "B"),
            )
            expected_formulas = {
                "F": self._day_normalized_yoy_formula(row, f"B{prior_row}")
            }
            blank = ()
        else:
            bindings = self._bindings_for_identity(formula_workbook, identity)
            persisted_columns = {
                config.metric_binding_by_id[item.metric_variant_id].column
                for item in bindings
            }
            quarter_transition = not {"H", "K", "L"}.issubset(persisted_columns)
            expected_formulas = {
                "F": self._day_normalized_yoy_formula(row, f"D{row}"),
            }
            if quarter_transition:
                expected_formulas.update(
                    {
                        "G": f"=B{row}",
                        "I": f"=D{row}",
                        "J": f"=E{row}",
                    }
                )
                blank = ("H", "K", "L")
            else:
                reporting_dates = {item.workflow_reporting_date for item in bindings}
                if len(reporting_dates) != 1:
                    return False
                reporting_date = date.fromisoformat(next(iter(reporting_dates)))
                previous_row = self._exact_reporting_row(
                    formula_workbook,
                    config,
                    (reporting_date - timedelta(days=7)).isoformat(),
                )
                expected_formulas.update(
                    {
                        "G": f"=B{row}-B{previous_row}",
                        "H": f"=C{row}-C{previous_row}",
                        "I": f"=D{row}-D{previous_row}",
                        "J": f"=E{row}-E{previous_row}",
                        "K": f"=H{row}/H{previous_row}-1",
                        "L": f"=H{row}/J{row}-1",
                    }
                )
                blank = ()
        for column, expected_formula in expected_formulas.items():
            formula = formula_sheet[f"{column}{row}"].value
            cached = value_sheet[f"{column}{row}"].value
            if formula != expected_formula:
                return False
            try:
                _decimal_value(cached, cell_reference=f"{column}{row}")
            except MetricStoreError:
                return False
        return all(
            formula_sheet[f"{column}{row}"].value is None
            and value_sheet[f"{column}{row}"].value is None
            for column in blank
        )

    def _exact_reporting_row(
        self,
        workbook: Workbook,
        config: RevenueExcelStoreAssetConfig,
        workflow_reporting_date: str,
    ) -> int:
        try:
            bindings = self._lineage.bindings_in_workbook(workbook)
        except MetricStoreError as exc:
            raise MetricStoreError(
                "STORE_EXACT_KEY_NOT_FOUND",
                "Exact previous reporting-period Store lineage is unavailable",
            ) from exc
        rows = {
            item.physical_row
            for item in bindings
            if item.store_id == config.store_id
            and item.store_asset_id == config.store_asset_id
            and item.business_context_id == config.business_context_id
            and item.workflow_reporting_date == workflow_reporting_date
            and item.validation_status == "passed"
        }
        if len(rows) != 1:
            code = "STORE_EXACT_KEY_NOT_FOUND" if not rows else "STORE_EXACT_KEY_AMBIGUOUS"
            raise MetricStoreError(
                code,
                "Previous reporting-period Store lineage must resolve to one physical row",
            )
        return next(iter(rows))

    def _exact_business_row(
        self,
        workbook: Workbook,
        config: RevenueExcelStoreAssetConfig,
        business_date: str,
        metric_variant_id: str,
    ) -> int:
        rows = self._physical_rows_for_date(workbook, config, business_date)
        if len(rows) != 1:
            code = (
                "STORE_EXACT_BUSINESS_DATE_NOT_FOUND"
                if not rows
                else "STORE_EXACT_BUSINESS_DATE_AMBIGUOUS"
            )
            raise MetricStoreError(
                code,
                "Prior-year business date must resolve to one physical Store row",
            )
        key = StoreBusinessDateReadKey(
            config.store_id,
            config.store_asset_id,
            metric_variant_id,
            business_date,
            config.business_context_id,
        )
        try:
            binding = self._lineage.read_exact_business_date(
                config.workbook_path, key
            )
            self._verified_record_group(config, binding, metric_variant_id)
        except MetricStoreError as exc:
            raise MetricStoreError(
                "STORE_EXACT_BUSINESS_DATE_NOT_VERIFIED",
                "Prior-year physical row has no verified technical lineage",
            ) from exc
        if binding.physical_row != rows[0]:
            raise MetricStoreError(
                "STORE_EXACT_BUSINESS_DATE_NOT_VERIFIED",
                "Prior-year physical row does not have one verified lineage binding",
            )
        return rows[0]

    @staticmethod
    def _metric_variant_for_column(
        config: RevenueExcelStoreAssetConfig, column: str
    ) -> str:
        matches = [
            item.metric_variant_id
            for item in config.metric_bindings
            if item.column == column
        ]
        if len(matches) != 1:
            raise MetricStoreError(
                "STORE_EXCEL_METRIC_BINDING_MISSING",
                f"Formula profile requires one Metric Variant in column {column}",
            )
        return matches[0]

    @staticmethod
    def _prior_year_business_date(current_date: str) -> str:
        current = date.fromisoformat(current_date)
        try:
            previous = current.replace(year=current.year - 1)
        except ValueError as exc:
            raise MetricStoreError(
                "STORE_EXCEL_PRIOR_YEAR_DATE_INVALID",
                "Prior-year business date cannot be derived exactly",
            ) from exc
        return (previous + timedelta(days=1)).isoformat()

    @staticmethod
    def _day_normalized_yoy_formula(row: int, denominator_cell: str) -> str:
        current = f"A{row}"
        prior = f"(DATE(YEAR({current})-1,MONTH({current}),DAY({current}))+1)"
        current_start = (
            f"DATE(YEAR({current}),INT((MONTH({current})-1)/3)*3+1,1)"
        )
        prior_start = f"DATE(YEAR({prior}),INT((MONTH({prior})-1)/3)*3+1,1)"
        return (
            f"=(B{row}/({current}-{current_start}+1))/"
            f"({denominator_cell}/({prior}-{prior_start}+1))-1"
        )

    @staticmethod
    def _copy_previous_row_style(worksheet, target_row: int, config) -> None:
        source_row = next(
            (
                row
                for row in range(target_row - 1, 1, -1)
                if worksheet[f"{config.business_date_column}{row}"].value is not None
            ),
            1,
        )
        if source_row < 2:
            return
        columns = {config.business_date_column} | {
            item.column for item in config.metric_bindings
        } | set(config.formula_columns)
        if config.formula_profile == "technical_full":
            columns |= {"D", "E", "G", "I", "J"}
        elif config.formula_profile == "ctv_yoy":
            columns |= {"D", "E"}
        for column in columns:
            source = worksheet[f"{column}{source_row}"]
            target = worksheet[f"{column}{target_row}"]
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format

    @staticmethod
    def _receipt(
        plan: StoreWritePlan, records: tuple[MetricStoreRecord, ...]
    ) -> StoreWriteReceipt:
        ordered = tuple(sorted(records, key=lambda item: item.metric_variant_id))
        return StoreWriteReceipt(
            plan.write_identity,
            tuple(item.read_key for item in ordered),
            tuple(item.result_id for item in ordered),
            plan.business_digest,
            plan.idempotent_replay,
        )
