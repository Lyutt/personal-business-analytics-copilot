"""Adapter-owned physical lineage metadata for Revenue Excel Metric Stores."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Final

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from .errors import MetricStoreError
from .store import StoreBusinessDateReadKey, StoreReadKey

METADATA_WORKSHEET: Final = "_pbac_metric_store_metadata"
SCHEMA_VERSION: Final = "1.0.0"
METADATA_COLUMNS: Final = (
    "schema_version",
    "store_id",
    "store_asset_id",
    "business_context_id",
    "metric_variant_id",
    "workflow_reporting_date",
    "current_revenue_cutoff_date",
    "physical_worksheet",
    "physical_row",
    "business_date_column",
    "result_id",
    "validation_status",
    "business_digest",
)


@dataclass(frozen=True)
class ExcelLineageBinding:
    """Technical metadata joining one logical Metric Result to one workbook row."""

    store_id: str
    store_asset_id: str
    business_context_id: str
    metric_variant_id: str
    workflow_reporting_date: str
    current_revenue_cutoff_date: str
    physical_worksheet: str
    physical_row: int
    business_date_column: str
    result_id: str
    validation_status: str
    business_digest: str

    @property
    def reporting_read_key(self) -> StoreReadKey:
        return StoreReadKey(
            self.store_id,
            self.store_asset_id,
            self.metric_variant_id,
            self.workflow_reporting_date,
            self.business_context_id,
        )

    @property
    def business_date_read_key(self) -> StoreBusinessDateReadKey:
        return StoreBusinessDateReadKey(
            self.store_id,
            self.store_asset_id,
            self.metric_variant_id,
            self.current_revenue_cutoff_date,
            self.business_context_id,
        )


def _canonical_date(value: object, *, field_name: str) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if len(text) == 8 and text.isdigit():
        text = f"{text[:4]}-{text[4:6]}-{text[6:]}"
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError as exc:
        raise MetricStoreError(
            "STORE_EXCEL_LINEAGE_DATE_INVALID",
            f"{field_name} is not an exact ISO or YYYYMMDD business date",
        ) from exc


class RevenueExcelLineageAdapter:
    """Persist and verify technical key-to-row bindings inside a Store workbook."""

    def register(self, workbook_path: Path, binding: ExcelLineageBinding) -> bool:
        workbook = load_workbook(workbook_path)
        changed = self.register_in_workbook(workbook, binding)
        if not changed:
            return False
        workbook.save(workbook_path)
        verified = self.read_exact(workbook_path, binding.reporting_read_key)
        if verified != binding:
            raise MetricStoreError(
                "STORE_EXCEL_LINEAGE_POST_WRITE_VERIFICATION_FAILED",
                "Saved technical lineage metadata did not verify after workbook reopen",
            )
        return True

    def register_in_workbook(
        self, workbook: Workbook, binding: ExcelLineageBinding
    ) -> bool:
        """Register one binding without saving, for an atomic business-row transaction."""

        self._validate_binding(workbook, binding)
        worksheet = self._metadata_worksheet(workbook, create=True)
        matches = [
            item
            for item in self._bindings(worksheet)
            if item.reporting_read_key == binding.reporting_read_key
        ]
        if matches:
            if len(matches) == 1 and matches[0] == binding:
                return False
            raise MetricStoreError(
                "STORE_EXCEL_LINEAGE_REPORTING_KEY_CONFLICT",
                "The exact reporting-date Store key is already bound differently",
            )
        worksheet.append((SCHEMA_VERSION, *self._binding_values(binding)))
        worksheet.sheet_state = "veryHidden"
        return True

    def bindings_in_workbook(self, workbook: Workbook) -> tuple[ExcelLineageBinding, ...]:
        """Return validated technical bindings without treating metadata as business data."""

        worksheet = self._metadata_worksheet(workbook, create=False)
        bindings = self._bindings(worksheet)
        for binding in bindings:
            self._validate_binding(workbook, binding)
        return bindings

    def read_exact(self, workbook_path: Path, key: StoreReadKey) -> ExcelLineageBinding:
        workbook = load_workbook(workbook_path, data_only=False, read_only=False)
        worksheet = self._metadata_worksheet(workbook, create=False)
        matches = [item for item in self._bindings(worksheet) if item.reporting_read_key == key]
        binding = self._require_one(matches, "STORE_EXCEL_LINEAGE_REPORTING_KEY")
        self._validate_binding(workbook, binding)
        return binding

    def read_exact_business_date(
        self, workbook_path: Path, key: StoreBusinessDateReadKey
    ) -> ExcelLineageBinding:
        workbook = load_workbook(workbook_path, data_only=False, read_only=False)
        worksheet = self._metadata_worksheet(workbook, create=False)
        matches = [
            item for item in self._bindings(worksheet) if item.business_date_read_key == key
        ]
        binding = self._require_one(matches, "STORE_EXCEL_LINEAGE_BUSINESS_DATE_KEY")
        self._validate_binding(workbook, binding)
        return binding

    @staticmethod
    def _require_one(
        matches: list[ExcelLineageBinding], code_prefix: str
    ) -> ExcelLineageBinding:
        if not matches:
            raise MetricStoreError(f"{code_prefix}_NOT_FOUND", "No exact metadata binding exists")
        if len(matches) != 1:
            raise MetricStoreError(
                f"{code_prefix}_AMBIGUOUS", "More than one exact metadata binding exists"
            )
        return matches[0]

    @staticmethod
    def _metadata_worksheet(workbook: Workbook, *, create: bool):
        if METADATA_WORKSHEET not in workbook.sheetnames:
            if not create:
                raise MetricStoreError(
                    "STORE_EXCEL_LINEAGE_METADATA_MISSING",
                    "Adapter technical metadata worksheet does not exist",
                )
            worksheet = workbook.create_sheet(METADATA_WORKSHEET)
            worksheet.append(METADATA_COLUMNS)
            worksheet.sheet_state = "veryHidden"
            return worksheet
        worksheet = workbook[METADATA_WORKSHEET]
        headers = tuple(cell.value for cell in worksheet[1])
        if headers != METADATA_COLUMNS:
            raise MetricStoreError(
                "STORE_EXCEL_LINEAGE_SCHEMA_INVALID",
                "Adapter technical metadata worksheet columns do not match the registered schema",
            )
        if worksheet.sheet_state != "veryHidden":
            raise MetricStoreError(
                "STORE_EXCEL_LINEAGE_VISIBILITY_INVALID",
                "Adapter technical metadata worksheet must remain veryHidden",
            )
        return worksheet

    @staticmethod
    def _binding_values(binding: ExcelLineageBinding) -> tuple[object, ...]:
        return (
            binding.store_id,
            binding.store_asset_id,
            binding.business_context_id,
            binding.metric_variant_id,
            _canonical_date(
                binding.workflow_reporting_date, field_name="workflow_reporting_date"
            ),
            _canonical_date(
                binding.current_revenue_cutoff_date,
                field_name="current_revenue_cutoff_date",
            ),
            binding.physical_worksheet,
            binding.physical_row,
            binding.business_date_column,
            binding.result_id,
            binding.validation_status,
            binding.business_digest,
        )

    def _bindings(self, worksheet) -> tuple[ExcelLineageBinding, ...]:
        results: list[ExcelLineageBinding] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None for value in values):
                continue
            if len(values) != len(METADATA_COLUMNS) or values[0] != SCHEMA_VERSION:
                raise MetricStoreError(
                    "STORE_EXCEL_LINEAGE_SCHEMA_INVALID",
                    "Adapter technical metadata row does not match the registered schema",
                )
            results.append(
                ExcelLineageBinding(
                    store_id=str(values[1]),
                    store_asset_id=str(values[2]),
                    business_context_id=str(values[3]),
                    metric_variant_id=str(values[4]),
                    workflow_reporting_date=_canonical_date(
                        values[5], field_name="workflow_reporting_date"
                    ),
                    current_revenue_cutoff_date=_canonical_date(
                        values[6], field_name="current_revenue_cutoff_date"
                    ),
                    physical_worksheet=str(values[7]),
                    physical_row=int(values[8]),
                    business_date_column=str(values[9]),
                    result_id=str(values[10]),
                    validation_status=str(values[11]),
                    business_digest=str(values[12]),
                )
            )
        return tuple(results)

    def _validate_binding(self, workbook: Workbook, binding: ExcelLineageBinding) -> None:
        if binding.validation_status != "passed":
            raise MetricStoreError(
                "STORE_EXCEL_LINEAGE_NOT_VALIDATED",
                "Only passed Metric Result lineage may be registered",
            )
        if binding.physical_worksheet == METADATA_WORKSHEET:
            raise MetricStoreError(
                "STORE_EXCEL_LINEAGE_PHYSICAL_TARGET_INVALID",
                "Technical metadata cannot be its own physical business target",
            )
        if binding.physical_worksheet not in workbook.sheetnames or binding.physical_row < 2:
            raise MetricStoreError(
                "STORE_EXCEL_LINEAGE_PHYSICAL_TARGET_INVALID",
                "The bound physical business row does not exist",
            )
        worksheet = workbook[binding.physical_worksheet]
        physical_date = _canonical_date(
            worksheet[f"{binding.business_date_column}{binding.physical_row}"].value,
            field_name="physical business-date cell",
        )
        requested_date = _canonical_date(
            binding.current_revenue_cutoff_date,
            field_name="current_revenue_cutoff_date",
        )
        if physical_date != requested_date:
            raise MetricStoreError(
                "STORE_EXCEL_LINEAGE_BUSINESS_DATE_MISMATCH",
                "The selected metadata cutoff does not equal the physical business-date cell",
            )
